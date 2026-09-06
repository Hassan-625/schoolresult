import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path
from django.test import TestCase, override_settings
from openpyxl import load_workbook
from .excel import compile_class
from .models import Result, School, Student, Subject

class CompilationTests(TestCase):
    def setUp(self): self.school=School.objects.get(slug="highflyers")
    def student(self, reg_no, class_level):
        return Student.objects.create(school=self.school,first_name="Test",last_name=reg_no,reg_no=reg_no,class_level=class_level,term="Third",session="2026/2027",term_ending=date(2027,7,20),next_term_begins=date(2027,9,10))
    def result(self, student, subject, first_ca, second_ca, exam):
        return Result.objects.create(school=self.school,student=student,subject=subject,first_ca=first_ca,second_ca=second_ca,exam=exam,first_term=60,second_term=70)
    def test_grade_duplicate_and_tie_ranking(self):
        subject=Subject.objects.get(school=self.school,name="MATHEMATICS",section="primary")
        a=self.student("A1","Basic 3"); b=self.student("B1","Basic 3")
        ra=self.result(a,subject,15,16,45); self.result(b,subject,15,16,45)
        self.assertEqual(ra.total,Decimal("76")); self.assertEqual(ra.grade,"A"); self.assertEqual(ra.remark,"Excellent")
        with self.assertRaises(Exception): self.result(a,subject,1,1,1)
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                compilation=compile_class(self.school,"Basic 3")
                a.refresh_from_db(); b.refresh_from_db()
                self.assertEqual((a.position,b.position),(1,1)); self.assertEqual(a.class_size,2)
                self.assertTrue(Path(compilation.zip_file.path).exists())
                wb=load_workbook(a.compiled_report.path)
                self.assertEqual(wb.sheetnames,["Sheet1","Sheet2","Sheet3"])
                self.assertEqual(wb["Sheet1"]["C9"].value,"A1"); self.assertEqual(wb["Sheet1"]["G13"].value,76)
    def test_nursery_template_mapping(self):
        subject=Subject.objects.get(school=self.school,name="MATHEMATICS",section="nursery")
        student=self.student("N1","Nursery 1"); self.result(student,subject,10,10,30)
        with tempfile.TemporaryDirectory() as media:
            with override_settings(MEDIA_ROOT=media):
                compile_class(self.school,"Nursery 1"); student.refresh_from_db()
                wb=load_workbook(student.compiled_report.path); ws=wb["Sheet1"]
                self.assertEqual(ws["G13"].value,50); self.assertEqual(ws["C27"].value,1); self.assertEqual(ws["H28"].value,"D")
