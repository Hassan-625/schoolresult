from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from django.conf import settings
from django.core.files import File
from django.utils import timezone
from openpyxl import load_workbook
from .models import Compilation, Student
from .services import compute_class_results, q, safe_name

MAPPING = {
 "primary": {"template":"Primary_section_template.xlsx", "start":13, "end":27, "grand":28, "size":"C30", "position":"C31", "average":"H30", "grade":"H31"},
 "nursery": {"template":"Nursery_section_template.xlsx", "start":13, "end":24, "grand":25, "size":"C27", "position":"C28", "average":"H27", "grade":"H28"},
}
def report_directory(student):
    return Path(settings.MEDIA_ROOT) / "results" / safe_name(student.session) / student.term / safe_name(student.class_level)
def generate_report_card(student):
    spec=MAPPING[student.section]; template=Path(__file__).resolve().parent / "templates" / "excel" / spec["template"]
    wb=load_workbook(template); ws=wb["Sheet1"]
    for cell,value in {"E7":student.term,"J7":student.session,"C8":student.full_name,"C9":student.reg_no,"C10":student.term_ending,"J9":student.class_level,"L10":student.next_term_begins}.items(): ws[cell]=value
    results={r.subject.name.strip().upper():r for r in student.results.select_related("subject")}
    for row in range(spec["start"], spec["end"]+1):
        result=results.get(str(ws[f"B{row}"].value or "").strip().upper())
        values = [result.first_ca,result.second_ca,result.exam,result.total,result.first_term,result.second_term,result.subject_average,result.grade,result.subject_position,result.class_average,result.remark] if result else [None]*11
        for col,value in zip("DEFGHIJKLMN", values): ws[f"{col}{row}"]=value
    rows=list(student.results.all()); ws[f"G{spec['grand']}"]=student.total_score
    ws[f"H{spec['grand']}"]=q(sum((r.first_term for r in rows),0)); ws[f"I{spec['grand']}"]=q(sum((r.second_term for r in rows),0))
    ws[spec["size"]]=student.class_size; ws[spec["position"]]=student.position; ws[spec["average"]]=student.average_score; ws[spec["grade"]]=student.overall_grade
    out_dir=report_directory(student); out_dir.mkdir(parents=True, exist_ok=True)
    out=out_dir / f"{safe_name(student.reg_no)}_{safe_name(student.full_name)}.xlsx"; wb.save(out)
    relative=out.relative_to(settings.MEDIA_ROOT).as_posix(); Student.objects.filter(pk=student.pk).update(compiled_report=relative); student.compiled_report.name=relative
    return out
def compile_class(class_level, term=None, session=None):
    students=compute_class_results(class_level, term, session)
    if not students: return None
    paths=[generate_report_card(s) for s in students]; first=students[0]
    out_dir=report_directory(first); zip_path=out_dir / f"{safe_name(class_level)}_{safe_name(first.term)}_{safe_name(first.session)}.zip"
    with ZipFile(zip_path,"w",ZIP_DEFLATED) as archive:
        for path in paths: archive.write(path, path.name)
    with zip_path.open("rb") as fh:
        obj=Compilation(class_level=class_level,term=first.term,session=first.session); obj.zip_file.save(zip_path.name,File(fh),save=True)
    return obj
